import csv
import io
from openpyxl import load_workbook

XLSX = r"C:\Users\droig\Proyectos\CSJ\buzon\entrante\Análisis Global_v2 (1).xlsx"
CSV  = r"C:\Users\droig\Proyectos\CSJ\buzon\entrante\cancelaciones.csv"
OUT  = r"C:\Users\droig\Proyectos\CSJ\buzon\entrante\_temp_out.txt"

MESES = {
    'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
    'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12
}

def parse_double_quoted_csv(filepath):
    """The CSV has each row wrapped in outer quotes — parse in two steps."""
    rows = []
    with open(filepath, newline='', encoding='utf-8') as f:
        outer = csv.reader(f)  # each "row" = [single_string_containing_commas]
        raw_rows = [r[0] for r in outer if r]
    if not raw_rows:
        return []
    # Parse header
    headers = next(csv.reader([raw_rows[0]]))
    headers = [h.strip().lstrip('\ufeff') for h in headers]
    # Parse data rows
    for raw in raw_rows[1:]:
        values = next(csv.reader([raw]))
        rows.append(dict(zip(headers, values)))
    return rows

# ── Leer GS ──────────────────────────────────────────────────────────────────
wb = load_workbook(XLSX, data_only=True)
gs_entries = []

for year in range(2015, 2027):
    sname = str(year)
    if sname not in wb.sheetnames:
        continue
    ws = wb[sname]
    mes_actual = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        col_d = row[3]
        if col_d and str(col_d).strip().upper() in MESES:
            mes_actual = MESES[str(col_d).strip().upper()]
        nombre = row[2]   # Col C
        nights = row[4]   # Col E
        cleaning = row[7] # Col H
        total = row[8]    # Col I
        pm = row[6]       # Col G
        if nombre and nights and isinstance(nights, (int, float)) and nights > 0:
            nombre_str = str(nombre).strip()
            first = nombre_str.split()[0].lower()
            gs_entries.append({
                'year': year, 'mes': mes_actual,
                'nombre': nombre_str, 'first': first,
                'nights': int(nights),
                'cleaning': cleaning, 'total': total, 'pm': pm,
            })

# ── Leer cancelaciones.csv ───────────────────────────────────────────────────
raw_rows = parse_double_quoted_csv(CSV)

cancels = []
for row in raw_rows:
    code       = row.get('Código de confirmación', '').strip()
    nombre     = row.get('Nombre de la persona', '').strip()
    noches_str = row.get('N.º de noches', '0').strip()
    fecha_ini  = row.get('Fecha de inicio', '').strip()
    ingresos   = row.get('Ingresos', '').replace('\xa0','').replace('€','').replace(',','.').strip()

    try: noches = int(noches_str)
    except: noches = 0

    year_c = mes_c = None
    if fecha_ini:
        parts = fecha_ini.split('/')
        if len(parts) == 3:
            try:
                year_c = int(parts[2])
                mes_c  = int(parts[1])
            except: pass

    first_c = nombre.split()[0].lower() if nombre else ''
    cancels.append({
        'code': code, 'nombre': nombre, 'first': first_c,
        'noches': noches, 'year': year_c, 'mes': mes_c,
        'ingresos': ingresos,
    })

# ── Matching ─────────────────────────────────────────────────────────────────
matched = []
multi   = []
no_match = []

for c in cancels:
    candidates = []
    for g in gs_entries:
        if g['year'] != c['year']:
            continue
        if g['first'] != c['first']:
            continue
        if abs(g['nights'] - c['noches']) > 1:
            continue
        if c['mes'] and g['mes'] and abs(g['mes'] - c['mes']) > 1:
            continue
        candidates.append(g)

    if len(candidates) == 1:
        matched.append((c, candidates[0]))
    elif len(candidates) > 1:
        multi.append((c, candidates))
    else:
        no_match.append(c)

# ── Output ───────────────────────────────────────────────────────────────────
lines = []
lines.append(f"=== MATCHED ({len(matched)}) ===")
for c, g in matched:
    mes_str = f"{c['mes']:02d}" if c['mes'] else '??'
    lines.append(f"{c['code']} | {c['nombre']} | {c['noches']}n | {c['year']}-{mes_str} | GS: {g['nombre']} | total={g['total']} | pm={g['pm']} | cleaning={g['cleaning']}")

lines.append(f"\n=== MÚLTIPLES CANDIDATOS ({len(multi)}) ===")
for c, cands in multi:
    mes_str = f"{c['mes']:02d}" if c['mes'] else '??'
    lines.append(f"{c['code']} | {c['nombre']} | {c['noches']}n | {c['year']}-{mes_str}")
    for g in cands:
        lines.append(f"   → {g['nombre']} | {g['nights']}n | mes={g['mes']} | total={g['total']} | pm={g['pm']}")

lines.append(f"\n=== SIN MATCH ({len(no_match)}) ===")
for c in no_match:
    lines.append(f"{c['code']} | {c['nombre']} | {c['noches']}n | {c['year']}-{c['mes']}")

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Total: {len(cancels)}, matched={len(matched)}, multi={len(multi)}, no_match={len(no_match)}")
