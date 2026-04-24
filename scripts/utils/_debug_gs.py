from openpyxl import load_workbook

XLSX = r"C:\Users\droig\Proyectos\CSJ\buzon\entrante\Análisis Global_v2 (1).xlsx"
OUT  = r"C:\Users\droig\Proyectos\CSJ\buzon\entrante\_temp_out.txt"

wb = load_workbook(XLSX, data_only=True)

lines = []
# Show first 60 rows of 2024 sheet to understand structure
ws = wb['2024']
lines.append("=== HOJA 2024 — primeras 80 filas ===")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), 1):
    if any(v is not None for v in row):
        lines.append(f"R{i:03d}: {list(row[:12])}")

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print("Done")
