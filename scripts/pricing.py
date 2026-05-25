#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TODAY = date.today()
DEFAULT_END   = date(DEFAULT_TODAY.year, 12, 31)

_ROOT          = Path(__file__).parent.parent
_DEFAULT_INPUT = str(_ROOT / "datos" / "reservas.json")
_DEFAULT_XLSX  = str(_ROOT / "output" / "pricing_output.xlsx")
_DEFAULT_JSON  = str(_ROOT / "output" / "pricing_output.json")

# ── Configuración general ──────────────────────────────────────────────────────
LOOKBACK_YEARS      = 5
MIN_VALID_PM        = 15.0
MAX_VALID_PM        = 500.0
WEEKLY_DISCOUNT     = 0.05    # descuento estancias >= 7 noches
NRF_DISCOUNT        = 0.10    # descuento tarifa NRF vs flexible
MAX_TENSION_P75     = 0.13    # tensión máxima sobre P75 en Rentabilidad (+13%)

# Zonas objetivo por estrategia (percentil inferior / superior)
STRATEGY_ZONES: Dict[str, Tuple[float, float]] = {
    "Ocupación":    (0.25, 0.40),
    "Equilibrio":   (0.45, 0.60),
    "Rentabilidad": (0.65, 0.75),
}

# ── Colores Excel ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
OCU_FILL    = PatternFill("solid", fgColor="FDE9D9")   # 🔴 Ocupación
EQU_FILL    = PatternFill("solid", fgColor="FFF2CC")   # 🟡 Equilibrio
RENT_FILL   = PatternFill("solid", fgColor="E8DAEF")   # 🟣 Rentabilidad
STRATEGY_FILL = {"Ocupación": OCU_FILL, "Equilibrio": EQU_FILL, "Rentabilidad": RENT_FILL}


# ── Dataclasses ────────────────────────────────────────────────────────────────

@dataclass
class Reservation:
    guest: str
    checkin: date
    checkout: date
    nights: int
    pm: float
    status: str
    booking_date: Optional[date]
    year: int
    month: int
    confirmation_code: str
    rate_type: Optional[str]

    @property
    def occupied_nights(self) -> List[date]:
        return [self.checkin + timedelta(days=i) for i in range(self.nights)]


@dataclass
class Gap:
    start_date: date
    end_date: date
    nights: int

    @property
    def month_key(self) -> str:
        return month_key_for_date(self.start_date)


@dataclass
class GapPricing:
    # Identificación
    start_date: str
    end_date: str
    nights: int
    month_bucket: str
    season: str
    lead_days: int
    future_occ_pct: float
    # Percentiles históricos
    p25: float
    p50: float
    p75: float
    # Estrategia y zona
    strategy: str
    zona_objetivo: str      # "P45–P60  (75€–92€)"
    zona_low: float
    zona_high: float
    # Precio
    precio_base_zona: float
    ajuste_rms: str
    recommended_price: float   # precio final flexible
    precio_nrf: float
    weekly_price: Optional[float]
    # Transparencia
    motivo_rms: str
    # Revisiones de precio
    review_date_30d: Optional[str]
    review_date_15d: Optional[str]
    review_date_7d: Optional[str]
    next_action: str
    # Límites de referencia
    floor_price: float
    ceiling_price: float


# ── Utilidades ─────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="RMS CSJ — pricing basado en zonas y estrategia multi-factor.")
    p.add_argument("--input",          default=_DEFAULT_INPUT)
    p.add_argument("--today",          default=DEFAULT_TODAY.isoformat())
    p.add_argument("--end",            default=DEFAULT_END.isoformat())
    p.add_argument("--pace",           type=float, default=None,
                   help="Factor de pace manual (0.80–1.20). Si no se indica, se calcula automáticamente.")
    p.add_argument("--lookback-years", type=int, default=LOOKBACK_YEARS)
    p.add_argument("--xlsx",           default=_DEFAULT_XLSX)
    p.add_argument("--json",           default=_DEFAULT_JSON)
    return p.parse_args()


def parse_date(value: str) -> Optional[date]:
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def month_key_for_date(d: date) -> str:
    return "12x" if (d.month == 12 and d.day >= 20) else f"{d.month:02d}"


def month_label(key: str) -> str:
    return {
        "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
        "05": "Mayo",  "06": "Junio",   "07": "Julio", "08": "Agosto",
        "09": "Septiembre", "10": "Octubre", "11": "Noviembre",
        "12": "Diciembre (1-19)", "12x": "Diciembre (20-31)",
    }.get(key, key)


def season_for_date(d: date) -> str:
    """Alta: 15 jun–15 sep. Media: abr–may, 1–14 jun, 16 sep–oct. Baja: nov–mar."""
    m, day = d.month, d.day
    if (m == 6 and day >= 15) or m in (7, 8) or (m == 9 and day <= 15):
        return "alta"
    if m in (4, 5) or (m == 6 and day < 15) or (m == 9 and day > 15) or m == 10:
        return "media"
    return "baja"


def percentile_linear(sv: List[float], p: float) -> float:
    if not sv:
        raise ValueError("Lista vacía para percentil.")
    if len(sv) == 1:
        return float(sv[0])
    pos  = (len(sv) - 1) * p
    low  = int(pos)
    high = min(low + 1, len(sv) - 1)
    return sv[low] * (1 - (pos - low)) + sv[high] * (pos - low)


# ── Carga y filtrado de reservas ───────────────────────────────────────────────

def load_reservations(path: Path) -> List[Reservation]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    result: List[Reservation] = []
    for row in data:
        checkin = parse_date(str(row.get("checkin", "")))
        nights  = int(row.get("nights", 0) or 0)
        if not checkin or nights <= 0:
            continue
        pm         = float(row.get("pm", 0) or 0)
        rate_type  = row.get("rate_type") or None
        pm_flex    = round(pm / (1 - NRF_DISCOUNT), 2) if rate_type == "nrf" and pm > 0 else pm
        result.append(Reservation(
            guest=str(row.get("guest", "")),
            checkin=checkin,
            checkout=checkin + timedelta(days=nights),
            nights=nights,
            pm=pm_flex,
            status=str(row.get("status", "confirmed")),
            booking_date=parse_date(str(row.get("booking_date", ""))),
            year=int(row.get("year", 0) or 0),
            month=int(row.get("month", 0) or 0),
            confirmation_code=str(row.get("confirmation_code", "")),
            rate_type=rate_type,
        ))
    return result


def filter_confirmed_with_valid_pm(reservations: List[Reservation], today: date, lookback: int) -> List[Reservation]:
    start_year = today.year - lookback
    return [
        r for r in reservations
        if r.status == "confirmed"
        and start_year <= r.checkin.year <= today.year
        and MIN_VALID_PM <= r.pm <= MAX_VALID_PM
        and r.nights < 28
    ]


# ── Percentiles y datos históricos ────────────────────────────────────────────

def compute_month_data(reservations: List[Reservation]) -> Tuple[Dict[str, Dict], Dict[str, List[float]]]:
    """Devuelve (percentiles_por_mes, valores_ordenados_por_mes)."""
    KEYS = ["01","02","03","04","05","06","07","08","09","10","11","12","12x"]
    buckets: Dict[str, List[float]] = {k: [] for k in KEYS}
    for r in reservations:
        buckets[month_key_for_date(r.checkin)].append(r.pm)

    percentiles: Dict[str, Dict] = {}
    sorted_vals: Dict[str, List[float]] = {}
    for key, vals in buckets.items():
        if not vals:
            continue
        sv = sorted(vals)
        sorted_vals[key] = sv
        percentiles[key] = {
            "count": float(len(sv)),
            "p25":   round(percentile_linear(sv, 0.25), 2),
            "p40":   round(percentile_linear(sv, 0.40), 2),
            "p45":   round(percentile_linear(sv, 0.45), 2),
            "p50":   round(percentile_linear(sv, 0.50), 2),
            "p60":   round(percentile_linear(sv, 0.60), 2),
            "p65":   round(percentile_linear(sv, 0.65), 2),
            "p75":   round(percentile_linear(sv, 0.75), 2),
            "min":   round(min(sv), 2),
            "max":   round(max(sv), 2),
        }
    return percentiles, sorted_vals


# ── Ocupación y huecos ─────────────────────────────────────────────────────────

def all_confirmed_nights(reservations: List[Reservation]) -> set:
    nights: set = set()
    for r in reservations:
        if r.status == "confirmed":
            for n in r.occupied_nights:
                nights.add(n)
    return nights


def detect_gaps(occupied: set, start: date, end: date) -> List[Gap]:
    gaps: List[Gap] = []
    current, gap_start = start, None
    while current <= end:
        free = current not in occupied
        if free and gap_start is None:
            gap_start = current
        elif not free and gap_start is not None:
            gaps.append(Gap(gap_start, current, (current - gap_start).days))
            gap_start = None
        current += timedelta(days=1)
    if gap_start is not None:
        final_end = end + timedelta(days=1)
        gaps.append(Gap(gap_start, final_end, (final_end - gap_start).days))
    return gaps


def compute_future_occ(gap_end: date, confirmed: set, horizon: int = 90) -> float:
    """% de noches ocupadas en los `horizon` días posteriores al fin del hueco."""
    occupied = sum(1 for i in range(horizon) if (gap_end + timedelta(days=i)) in confirmed)
    return round(occupied / horizon * 100, 1)


# ── Pace auto-calculado ────────────────────────────────────────────────────────

def compute_auto_pace(reservations: List[Reservation], today: date) -> float:
    """Compara revenue OTB año actual vs mismo punto del año anterior."""
    cy, py = today.year, today.year - 1
    same_day_py = date(py, today.month, min(today.day, 28))

    def otb_rev(year: int, cutoff: date) -> float:
        return sum(
            r.pm * r.nights for r in reservations
            if r.status == "confirmed" and r.year == year
            and r.booking_date and r.booking_date <= cutoff and r.pm > 0
        )

    cy_rev = otb_rev(cy, today)
    py_rev = otb_rev(py, same_day_py)
    if py_rev == 0:
        return 1.00
    return round(min(1.25, max(0.75, cy_rev / py_rev)), 3)


# ── Lógica de estrategia ───────────────────────────────────────────────────────

def choose_strategy(
    nights: int, lead_days: int, pace: float, season: str, future_occ_pct: float
) -> Tuple[str, List[str]]:
    """
    Determina estrategia y lista de motivos.

    Ocupación:    hueco corto (≤3n) + al menos un factor de presión.
    Rentabilidad: múltiples señales positivas simultáneas.
    Equilibrio:   estándar (default).
    """
    reasons: List[str] = []

    # ── Señales negativas (presión para vender) ────────────────────────────────
    is_short        = nights <= 3
    is_very_low_lead = lead_days < 10
    is_low_lead     = lead_days < 21
    is_weak_pickup  = pace < 0.97
    is_low_season   = season == "baja"
    is_low_occ      = future_occ_pct < 35

    # ── Señales positivas (contexto favorable) ────────────────────────────────
    is_high_season  = season == "alta"
    is_strong_pace  = pace >= 1.05
    is_good_lead    = lead_days >= 30
    is_high_occ     = future_occ_pct >= 60

    # ── Ocupación: hueco corto + presión ──────────────────────────────────────
    pressure = is_low_lead or is_weak_pickup or is_low_season or is_low_occ
    if is_short and pressure:
        reasons.append("hueco muy corto" if nights <= 2 else "hueco corto")
        if is_very_low_lead:   reasons.append("antelación muy baja")
        elif is_low_lead:      reasons.append("baja antelación")
        if is_weak_pickup:     reasons.append("pickup débil")
        if is_low_season:      reasons.append("temporada baja")
        if is_low_occ:         reasons.append("baja ocupación futura")
        return "Ocupación", reasons

    # ── Rentabilidad: mínimo 3 señales positivas o combo fuerte ───────────────
    pos_score = sum([is_high_season, is_strong_pace, is_good_lead, is_high_occ])
    strong_combo = is_high_season and is_good_lead and (is_strong_pace or is_high_occ)
    if pos_score >= 3 or strong_combo:
        if is_high_season:  reasons.append("temporada alta")
        if is_strong_pace:  reasons.append("pickup fuerte")
        if is_good_lead:    reasons.append("suficiente antelación")
        if is_high_occ:     reasons.append("ocupación futura alta")
        return "Rentabilidad", reasons

    # ── Equilibrio: modo estándar ──────────────────────────────────────────────
    reasons.append(f"temporada {season}")
    if is_good_lead:    reasons.append("antelación adecuada")
    elif is_low_lead:   reasons.append("antelación justa")
    return "Equilibrio", reasons


# ── Pricing dentro de la zona ──────────────────────────────────────────────────

def compute_price_in_zone(
    strategy: str,
    sorted_vals: List[float],
    pct: Dict[str, float],
    lead_days: int,
    pace: float,
    nights: int,
    future_occ_pct: float,
) -> Tuple[float, float, str, float, str]:
    """
    Devuelve: (zona_low_eur, zona_high_eur, zona_str, precio_final, ajuste_desc)

    Posición dentro de la zona [0.0 = borde inferior, 1.0 = borde superior].
    Parte del punto medio (0.5) y aplica microajustes según contexto.
    """
    p_low_q, p_high_q = STRATEGY_ZONES[strategy]
    p25    = pct["p25"]
    p75    = pct["p75"]
    p_low  = percentile_linear(sorted_vals, p_low_q)
    p_high = percentile_linear(sorted_vals, p_high_q)
    zona_str   = f"P{int(p_low_q*100)}–P{int(p_high_q*100)}  ({p_low:.0f}€–{p_high:.0f}€)"
    zone_range = max(p_high - p_low, 1.0)

    position   = 0.5        # midpoint de la zona
    adj_parts: List[str] = []

    # Antelación (lead time)
    if lead_days < 10:
        position -= 0.20; adj_parts.append("antelación muy baja (↓20%)")
    elif lead_days < 21:
        position -= 0.10; adj_parts.append("baja antelación (↓10%)")
    elif lead_days > 60:
        position += 0.15; adj_parts.append("buena antelación (↑15%)")
    elif lead_days > 30:
        position += 0.05

    # Pickup / pace
    if pace > 1.07:
        position += 0.20; adj_parts.append("pickup muy fuerte (↑20%)")
    elif pace > 1.02:
        position += 0.10; adj_parts.append("pickup fuerte (↑10%)")
    elif pace < 0.93:
        position -= 0.15; adj_parts.append("pickup débil (↓15%)")
    elif pace < 0.97:
        position -= 0.08; adj_parts.append("pickup bajo (↓8%)")

    # Duración del hueco
    if nights <= 2:
        position -= 0.10; adj_parts.append("hueco muy corto (↓10%)")
    elif nights <= 3:
        position -= 0.05; adj_parts.append("hueco corto (↓5%)")

    # Ocupación futura
    if future_occ_pct >= 70:
        position += 0.10; adj_parts.append("ocupación futura alta (↑10%)")
    elif future_occ_pct < 30:
        position -= 0.10; adj_parts.append("baja ocupación futura (↓10%)")

    position  = max(0.0, min(1.0, position))
    base_price = p_low + zone_range * position

    # Tensión sobre P75 (exclusivo Rentabilidad, con justificación explícita)
    tension_str = ""
    if strategy == "Rentabilidad":
        t_score = sum([lead_days > 45, pace > 1.05, future_occ_pct >= 60, nights >= 5])
        if t_score >= 3:
            tension_pct = min(0.04 * (t_score - 2), MAX_TENSION_P75)
            tensioned   = p75 * (1 + tension_pct)
            if tensioned > base_price:
                base_price  = tensioned
                tension_str = f"tensión +{int(tension_pct*100)}% sobre P75"

    # Límites absolutos: nunca por debajo de P25 ni por encima de P75+13%
    floor   = p25 * (0.90 if nights == 1 else 1.00)
    ceiling = p75 * (1 + MAX_TENSION_P75)
    final   = max(floor, min(base_price, ceiling))

    parts = adj_parts + ([tension_str] if tension_str else [])
    adj_str = " + ".join(parts) if parts else "sin ajustes adicionales"

    return p_low, p_high, zona_str, round(final, 2), adj_str


def build_motivo(strategy: str, strategy_reasons: List[str], adj_str: str) -> str:
    reason_text = " + ".join(strategy_reasons) if strategy_reasons else "criterio estándar"
    if strategy == "Ocupación":
        base = f"Ocupación ({reason_text}). Objetivo: convertir rápido, evitar hueco vacío."
    elif strategy == "Rentabilidad":
        base = f"Rentabilidad ({reason_text}). Objetivo: defender ADR sin descuento innecesario."
    else:
        base = f"Equilibrio ({reason_text}). Precio de mercado balanceado."
    if adj_str and adj_str != "sin ajustes adicionales":
        return f"{base} Ajustes en zona: {adj_str}."
    return base


# ── Fechas de revisión ─────────────────────────────────────────────────────────

def compute_review_dates(today: date, start: date) -> Tuple[Optional[str], Optional[str], Optional[str], str]:
    d30 = start - timedelta(days=30)
    d15 = start - timedelta(days=15)
    d7  = start - timedelta(days=7)
    v30 = d30.isoformat() if d30 >= today else None
    v15 = d15.isoformat() if d15 >= today else None
    v7  = d7.isoformat()  if d7  >= today else None
    if v30:    action = f"Revisar {v30}; sin vender → bajar 5%"
    elif v15:  action = f"Revisar {v15}; sin vender → bajar 5–10%"
    elif v7:   action = f"Revisar {v7}; sin vender → precio Ocupación"
    else:      action = "Hueco próximo — revisar diariamente"
    return v30, v15, v7, action


# ── Motor principal de pricing ─────────────────────────────────────────────────

def price_gaps(
    gaps: List[Gap],
    percentiles: Dict[str, Dict],
    sorted_vals_by_key: Dict[str, List[float]],
    confirmed: set,
    today: date,
    pace: float,
) -> List[GapPricing]:
    results: List[GapPricing] = []
    for gap in gaps:
        key = gap.month_key
        if key not in percentiles:
            key = f"{gap.start_date.month:02d}"
            if key not in percentiles:
                continue

        pct = percentiles[key]
        sv  = sorted_vals_by_key[key]

        lead_days      = (gap.start_date - today).days
        season         = season_for_date(gap.start_date)
        future_occ     = compute_future_occ(gap.end_date, confirmed)

        strategy, s_reasons = choose_strategy(gap.nights, lead_days, pace, season, future_occ)
        zona_low, zona_high, zona_str, final_price, adj_str = compute_price_in_zone(
            strategy, sv, pct, lead_days, pace, gap.nights, future_occ
        )
        motivo = build_motivo(strategy, s_reasons, adj_str)

        p25 = pct["p25"]; p50 = pct["p50"]; p75 = pct["p75"]
        floor   = p25 * (0.90 if gap.nights == 1 else 1.00)
        ceiling = p75 * (1 + MAX_TENSION_P75)
        weekly  = round(final_price * 7 * (1 - WEEKLY_DISCOUNT), 2) if gap.nights >= 7 else None
        nrf     = round(final_price * (1 - NRF_DISCOUNT), 2)
        r30, r15, r7, action = compute_review_dates(today, gap.start_date)

        results.append(GapPricing(
            start_date=gap.start_date.isoformat(),
            end_date=gap.end_date.isoformat(),
            nights=gap.nights,
            month_bucket=month_label(key),
            season=season,
            lead_days=lead_days,
            future_occ_pct=future_occ,
            p25=p25, p50=p50, p75=p75,
            strategy=strategy,
            zona_objetivo=zona_str,
            zona_low=round(zona_low, 2),
            zona_high=round(zona_high, 2),
            precio_base_zona=round(zona_low + (zona_high - zona_low) * 0.5, 2),
            ajuste_rms=adj_str,
            recommended_price=final_price,
            precio_nrf=nrf,
            weekly_price=weekly,
            motivo_rms=motivo,
            review_date_30d=r30,
            review_date_15d=r15,
            review_date_7d=r7,
            next_action=action,
            floor_price=round(floor, 2),
            ceiling_price=round(ceiling, 2),
        ))
    results.sort(key=lambda x: x.start_date)
    return results


# ── Excel ──────────────────────────────────────────────────────────────────────

def autosize_worksheet(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)


def write_workbook(
    xlsx_path: Path,
    gaps: List[GapPricing],
    percentiles: Dict[str, Dict],
    pace: float,
    pace_manual: bool,
    pace_auto: float,
    log_row: Dict,
) -> None:
    wb = Workbook()

    # ── Hoja Huecos ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Huecos"
    headers = [
        "Inicio", "Fin", "Noches", "Mes", "Temporada", "Lead (días)", "Ocup. futura %",
        "P25", "P50", "P75",
        "Estrategia", "Zona objetivo",
        "Precio base zona", "Ajuste RMS", "Precio final (flex)",
        "Precio NRF (−10%)", "Precio 7n (−5%)",
        "Motivo RMS",
        "Revisión 30d", "Revisión 15d", "Revisión 7d", "Próxima acción",
        "Suelo (P25)", "Techo (P75+13%)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row in gaps:
        ws.append([
            row.start_date, row.end_date, row.nights, row.month_bucket,
            row.season, row.lead_days, row.future_occ_pct,
            row.p25, row.p50, row.p75,
            row.strategy, row.zona_objetivo,
            row.precio_base_zona, row.ajuste_rms, row.recommended_price,
            row.precio_nrf, row.weekly_price,
            row.motivo_rms,
            row.review_date_30d, row.review_date_15d, row.review_date_7d, row.next_action,
            row.floor_price, row.ceiling_price,
        ])

    for r in range(2, ws.max_row + 1):
        fill = STRATEGY_FILL.get(ws.cell(r, 11).value)
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill

    autosize_worksheet(ws)
    ws.freeze_panes = "A2"

    # ── Hoja Percentiles ──────────────────────────────────────────────────────
    wp = wb.create_sheet("Percentiles")
    wp.append(["Mes", "P25", "P40", "P45", "P50", "P60", "P65", "P75", "Min", "Max", "N"])
    for cell in wp[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for key in ["01","02","03","04","05","06","07","08","09","10","11","12","12x"]:
        if key in percentiles:
            p = percentiles[key]
            wp.append([
                month_label(key),
                p["p25"], p["p40"], p["p45"], p["p50"], p["p60"], p["p65"], p["p75"],
                p["min"], p["max"], int(p["count"])
            ])
    autosize_worksheet(wp)
    wp.freeze_panes = "A2"

    # ── Hoja Parámetros ───────────────────────────────────────────────────────
    wc = wb.create_sheet("Parametros")
    wc.append(["Parámetro", "Valor"])
    for cell in wc[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    pace_label = f"{pace:.3f} (manual)" if pace_manual else f"{pace:.3f} (auto — vs {pace_auto:.3f} auto)"
    for row in [
        ("Fecha cálculo",          log_row["today"]),
        ("Fecha final",            log_row["end"]),
        ("Lookback years",         log_row["lookback_years"]),
        ("Pace factor",            pace_label),
        ("Pace auto-calculado",    f"{pace_auto:.3f}"),
        ("Descuento semanal ≥7n",  f"{int(WEEKLY_DISCOUNT*100)}%"),
        ("Descuento NRF",          f"{int(NRF_DISCOUNT*100)}%"),
        ("Tensión máx. sobre P75", f"+{int(MAX_TENSION_P75*100)}%"),
        ("",                       ""),
        ("Zona Ocupación",         "P25–P40  →  objetivo: conversión"),
        ("Zona Equilibrio",        "P45–P60  →  balance estándar"),
        ("Zona Rentabilidad",      "P65–P75 (+tensión si contexto)  →  defender ADR"),
        ("",                       ""),
        ("Suelo absoluto",         "P25 (1 noche: P25×0.90)"),
        ("Techo absoluto",         f"P75×{1+MAX_TENSION_P75:.2f}"),
        ("",                       ""),
        ("Ocupación dispara si",   "hueco ≤3n + (lead<21d ó pace<0.97 ó baja temporada ó occ_fut<35%)"),
        ("Rentabilidad dispara si","≥3 señales positivas: alta temporada, pickup fuerte, lead≥30d, occ_fut≥60%"),
        ("Equilibrio",             "estándar (default cuando no hay señales dominantes)"),
    ]:
        wc.append(row)
    autosize_worksheet(wc)

    # ── Hoja Log ──────────────────────────────────────────────────────────────
    wl = wb.create_sheet("Log")
    wl.append(["Timestamp", "Today", "End", "Lookback", "Pace", "Pace_auto", "Huecos", "Notas"])
    for cell in wl[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    wl.append([
        log_row["timestamp"], log_row["today"], log_row["end"],
        log_row["lookback_years"], f"{pace:.3f}",
        f"{pace_auto:.3f}", log_row["gaps_detected"], log_row["notes"],
    ])
    autosize_worksheet(wl)

    wb.save(xlsx_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    xlsx_path  = Path(args.xlsx)
    json_path  = Path(args.json)

    today = parse_date(args.today) or date.today()
    end   = parse_date(args.end)   or date(today.year, 12, 31)
    if end < today:
        raise ValueError("La fecha final no puede ser anterior a today.")

    reservations = load_reservations(input_path)
    valid        = filter_confirmed_with_valid_pm(reservations, today, args.lookback_years)
    percentiles, sorted_vals = compute_month_data(valid)

    pace_auto   = compute_auto_pace(reservations, today)
    pace_manual = args.pace is not None
    pace        = args.pace if pace_manual else pace_auto

    confirmed   = all_confirmed_nights(reservations)
    gap_occupied = {n for n in confirmed if today <= n <= end}
    gaps        = detect_gaps(gap_occupied, today, end)
    priced      = price_gaps(gaps, percentiles, sorted_vals, confirmed, today, pace)

    with json_path.open("w", encoding="utf-8") as f:
        json.dump([asdict(x) for x in priced], f, ensure_ascii=False, indent=2)

    log_row = {
        "timestamp":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "today":         today.isoformat(),
        "end":           end.isoformat(),
        "lookback_years": str(args.lookback_years),
        "pace":          f"{pace:.3f}",
        "gaps_detected": str(len(priced)),
        "notes":         "RMS v2 — estrategias multi-factor, zonas P25-P75, ajustes dinámicos.",
    }
    write_workbook(xlsx_path, priced, percentiles, pace, pace_manual, pace_auto, log_row)

    print(f"Excel:  {xlsx_path.resolve()}")
    print(f"JSON:   {json_path.resolve()}")
    print(f"Pace:   {pace:.3f} ({'manual' if pace_manual else f'auto — {pace_auto:.3f}'})")
    print(f"Huecos: {len(priced)}\n")
    for row in priced:
        print(
            f"{row.start_date} -> {row.end_date} | {row.nights:2}n | "
            f"{row.strategy:12} | {row.recommended_price:6.2f}EUR | "
            f"{row.zona_objetivo}"
        )


if __name__ == "__main__":
    main()
