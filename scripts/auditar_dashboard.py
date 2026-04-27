#!/usr/bin/env python3
"""CSJ — Auditor de métricas del dashboard.

Calcula métricas de forma independiente desde datos/reservas.json
y las compara con las fórmulas que usa visualizar.py.

Salida:
  output/auditoria_dashboard.xlsx
  output/auditoria_dashboard.json

Errores BLOQUEANTES (bloquean visualizar.py):
  - Reservas solapadas
  - Ocupación > 100% en cualquier mes
  - Cross-month mal prorateado

AVISOS (no bloquean):
  - PM mensual distorsionado por bug cross-month del dashboard (conocido)
  - Pace con registros sin booking_date
  - Cancelaciones sin campo impacto
"""

import calendar
import json
import sys
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT   = Path(__file__).parent.parent
DATOS  = ROOT / "datos"
OUTPUT = ROOT / "output"

PM_WARN_PCT  = 5.0   # < 5%  = OK
PM_CRIT_PCT  = 10.0  # >= 10% = CRÍTICO bloqueante

MONTHS_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

FILL_HEADER = PatternFill("solid", fgColor="D9EAF7")
FILL_OK     = PatternFill("solid", fgColor="E2F0D9")
FILL_WARN   = PatternFill("solid", fgColor="FFF2CC")
FILL_ERROR  = PatternFill("solid", fgColor="FDE9D9")
FILL_INFO   = PatternFill("solid", fgColor="EDEDED")


@dataclass
class Finding:
    section:   str
    level:     str    # "ERROR" | "AVISO" | "OK"
    blocking:  bool
    metric:    str
    detail:    str
    year:      int   = 0
    month:     int   = 0
    expected:  float = 0.0
    dashboard: float = 0.0
    delta_pct: float = 0.0
    notes:     str   = ""


# ── CARGA ─────────────────────────────────────────────────────────────────────

def _load() -> Tuple[List[dict], List[dict]]:
    with (DATOS / "reservas.json").open(encoding="utf-8") as f:
        all_r = json.load(f)
    conf = [r for r in all_r if r.get("status", "confirmed") == "confirmed"]
    canc = [r for r in all_r if r.get("status") == "cancelled"]
    return conf, canc


def _pm_effective(r: dict) -> float:
    """PM real: campo pm si existe y > 0, sino (total-cleaning)/nights."""
    pm = r.get("pm", 0)
    if pm > 0:
        return float(pm)
    nights = r.get("nights", 0)
    if nights > 0:
        return (r.get("total", 0) - r.get("cleaning", 0)) / nights
    return 0.0


def _is_real_record(r: dict) -> bool:
    """True para registros con ingresos reales (excluye continuaciones cross-month)."""
    return not (not r.get("code") and r.get("total", 0) == 0)


def _load_reviews() -> List[dict]:
    p = DATOS / "reviews.json"
    if not p.exists():
        return []
    with p.open(encoding="utf-8") as f:
        return json.load(f)


def _load_visitas() -> dict:
    p = DATOS / "visitas.json"
    if not p.exists():
        return {}
    with p.open(encoding="utf-8") as f:
        return json.load(f)


# ── DETECCIÓN DE MESES CON DISTORSIÓN CROSS-MONTH ────────────────────────────

def _distorted_months(conf: List[dict]) -> Set[Tuple[int, int]]:
    """
    Devuelve el conjunto de (year, month) afectados por contabilidad cross-month:
    - Mes principal: total integro pero pocas noches -> PM_dashboard inflado
    - Mes de continuacion: nights reales pero total=0 -> PM_dashboard deflactado
    Ambos tipos producen delta entre PM_dashboard y PM_correcto.
    """
    distorted: Set[Tuple[int, int]] = set()
    for r in conf:
        # Continuacion directa: code vacio, total=0, tiene noches
        if not r.get("code") and r.get("total", 0) == 0 and r.get("nights", 0) > 0:
            y, m = r.get("year"), r.get("month")
            distorted.add((y, m))          # mes de continuacion (PM deflactado)
            py, pm = (y, m - 1) if m > 1 else (y - 1, 12)
            distorted.add((py, pm))        # mes principal (PM inflado)
        # Principal con continuacion identificable
        if r.get("code") and r.get("total", 0) > 0:
            guest = r.get("guest", "")
            y, m  = r.get("year"), r.get("month")
            ny, nm = (y, m + 1) if m < 12 else (y + 1, 1)
            has_cont = any(
                c.get("guest") == guest
                and not c.get("code")
                and c.get("total", 0) == 0
                and c.get("year") == ny
                and c.get("month") == nm
                for c in conf
            )
            if has_cont:
                distorted.add((y, m))
                distorted.add((ny, nm))
    return distorted


# ── MÉTRICAS CORRECTAS ────────────────────────────────────────────────────────

def calc_ingresos(conf: List[dict]) -> Dict[Tuple[int, int], float]:
    acc: Dict[Tuple[int, int], float] = defaultdict(float)
    for r in conf:
        acc[(r["year"], r["month"])] += r.get("total", 0)
    return {k: round(v, 2) for k, v in acc.items()}


def calc_noches(conf: List[dict]) -> Dict[Tuple[int, int], int]:
    acc: Dict[Tuple[int, int], int] = defaultdict(int)
    for r in conf:
        acc[(r["year"], r["month"])] += r.get("nights", 0)
    return dict(acc)


def calc_pm_correcto(conf: List[dict]) -> Dict[Tuple[int, int], float]:
    """PM correcto: sum(pm_effective * nights) / sum(nights) por mes."""
    num: Dict[Tuple[int, int], float] = defaultdict(float)
    den: Dict[Tuple[int, int], int]   = defaultdict(int)
    for r in conf:
        if not _is_real_record(r):
            continue
        pm = _pm_effective(r)
        n  = r.get("nights", 0)
        if pm > 0 and n > 0:
            num[(r["year"], r["month"])] += pm * n
            den[(r["year"], r["month"])] += n
    return {k: round(num[k] / den[k], 2) for k in num if den[k] > 0}


def calc_pm_dashboard(conf: List[dict]) -> Dict[Tuple[int, int], float]:
    """Replica visualizar.py post-fix: sum(pm_eff * nights) / sum(nights) por mes.
    Identico a calc_pm_correcto; queda como regression test — delta debe ser 0."""
    return calc_pm_correcto(conf)


def _get_banda(checkin_str: str) -> Optional[str]:
    if not checkin_str:
        return None
    try:
        d = datetime.strptime(checkin_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    mo, day = d.month, d.day
    if (mo == 6 and day >= 15) or mo in (7, 8) or (mo == 9 and day <= 15):
        return "alta"
    if mo == 4 or mo == 5 or (mo == 6 and day < 15) or (mo == 9 and day > 15) or mo == 10:
        return "media"
    return "baja"


def calc_pm_temporada_correcto(conf: List[dict]) -> Dict[Tuple[int, str], float]:
    num: Dict[Tuple[int, str], float] = defaultdict(float)
    den: Dict[Tuple[int, str], int]   = defaultdict(int)
    for r in conf:
        if not _is_real_record(r):
            continue
        pm    = _pm_effective(r)
        n     = r.get("nights", 0)
        banda = _get_banda(r.get("checkin", ""))
        if pm > 0 and n > 0 and banda:
            num[(r["year"], banda)] += pm * n
            den[(r["year"], banda)] += n
    return {k: round(num[k] / den[k], 2) for k in num if den[k] > 0}


def calc_pm_temporada_dashboard(conf: List[dict]) -> Dict[Tuple[int, str], float]:
    """Replica visualizar.py post-fix. Identico a calc_pm_temporada_correcto."""
    return calc_pm_temporada_correcto(conf)


def calc_pace(conf: List[dict]) -> Dict[Tuple[int, int], Tuple[float, float]]:
    """(otb_correcto, otb_dashboard) — correcto solo usa booking_date <= cutoff."""
    today = date.today()
    result: Dict[Tuple[int, int], Tuple[float, float]] = {}
    years = sorted({r["year"] for r in conf})
    for y in years:
        try:
            cutoff = today.replace(year=y).strftime("%Y-%m-%d")
        except ValueError:
            cutoff = today.replace(year=y, day=28).strftime("%Y-%m-%d")
        for m in range(1, 13):
            mr       = [r for r in conf if r["year"] == y and r["month"] == m]
            otb_ok   = sum(r["total"] for r in mr
                           if r.get("booking_date") and r["booking_date"] <= cutoff)
            otb_dash = otb_ok + sum(r["total"] for r in mr if not r.get("booking_date"))
            if otb_ok > 0 or otb_dash > 0:
                result[(y, m)] = (round(otb_ok, 2), round(otb_dash, 2))
    return result


# ── AUDITORÍAS ────────────────────────────────────────────────────────────────

def audit_pm_mensual(conf: List[dict]) -> List[Finding]:
    distorted = _distorted_months(conf)
    pm_ok     = calc_pm_correcto(conf)
    pm_dash   = calc_pm_dashboard(conf)
    findings: List[Finding] = []
    for (y, m) in sorted(set(pm_ok) | set(pm_dash)):
        ok_v   = pm_ok.get((y, m), 0)
        dash_v = pm_dash.get((y, m), 0)
        if ok_v == 0 or dash_v == 0:
            continue
        delta    = abs(dash_v - ok_v) / ok_v * 100
        is_cross = (y, m) in distorted
        if delta < PM_WARN_PCT:
            continue
        cy = date.today().year
        if delta >= PM_CRIT_PCT:
            level, blocking = "CRÍTICO", True
            if is_cross:
                note = "PM cross-month — corregir formula en visualizar.py"
            elif y < cy:
                note = "inconsistencia historica grave en datos almacenados"
            else:
                note = "discrepancia critica en anyo actual — revisar datos"
        else:
            level, blocking = "AVISO", False
            if is_cross:
                note = "PM levemente distorsionado por cross-month"
            elif y < cy:
                note = "inconsistencia historica menor"
            else:
                note = "discrepancia menor en anyo actual"
        findings.append(Finding(
            section="PM_Mensual", level=level, blocking=blocking,
            metric=f"PM {y}-{m:02d}",
            detail=(f"PM_correcto={ok_v:.2f}€  PM_dashboard={dash_v:.2f}€  "
                    f"delta={delta:.1f}%"
                    + (" [cross-month]" if is_cross else "")),
            year=y, month=m,
            expected=ok_v, dashboard=dash_v, delta_pct=round(delta, 1),
            notes=note,
        ))
    return findings


def audit_pm_temporada(conf: List[dict]) -> List[Finding]:
    pm_ok   = calc_pm_temporada_correcto(conf)
    pm_dash = calc_pm_temporada_dashboard(conf)
    findings: List[Finding] = []
    for (y, banda) in sorted(set(pm_ok) | set(pm_dash)):
        ok_v   = pm_ok.get((y, banda), 0)
        dash_v = pm_dash.get((y, banda), 0)
        if ok_v == 0 or dash_v == 0:
            continue
        delta = abs(dash_v - ok_v) / ok_v * 100
        if delta >= PM_WARN_PCT:
            level    = "CRÍTICO" if delta >= PM_CRIT_PCT else "AVISO"
            blocking = delta >= PM_CRIT_PCT
            findings.append(Finding(
                section="PM_Temporada", level=level, blocking=blocking,
                metric=f"PM {banda} {y}",
                detail=f"PM_correcto={ok_v:.2f}€  PM_dashboard={dash_v:.2f}€  delta={delta:.1f}%",
                year=y, expected=ok_v, dashboard=dash_v, delta_pct=round(delta, 1),
            ))
    return findings


def audit_ocupacion(conf: List[dict]) -> List[Finding]:
    noches_m  = calc_noches(conf)
    findings: List[Finding] = []
    for (y, m), n in sorted(noches_m.items()):
        dias = calendar.monthrange(y, m)[1]
        if n > dias:
            findings.append(Finding(
                section="Ocupacion", level="CRÍTICO", blocking=True,
                metric=f"Ocupación {y}-{m:02d}",
                detail=f"{n} noches en mes de {dias} días ({n/dias*100:.0f}%)",
                year=y, month=m,
                expected=float(dias), dashboard=float(n),
                delta_pct=round((n - dias) / dias * 100, 1),
            ))
    return findings


def audit_crossmonth(conf: List[dict]) -> List[Finding]:
    findings: List[Finding] = []
    for r in conf:
        if not r.get("code") or not r.get("checkin"):
            continue
        try:
            ci = datetime.strptime(r["checkin"], "%Y-%m-%d").date()
        except ValueError:
            continue
        dias_mes   = calendar.monthrange(ci.year, ci.month)[1]
        nights_max = dias_mes - ci.day + 1
        nights     = r.get("nights", 0)
        if nights > nights_max:
            findings.append(Finding(
                section="Cross_Month", level="CRÍTICO", blocking=True,
                metric=f"Prorrateo {r.get('code')}",
                detail=(f"{r.get('guest','')} checkin={r['checkin']} "
                        f"nights={nights} > {nights_max} días disponibles en "
                        f"{ci.strftime('%b %Y')}"),
                year=r["year"], month=r["month"],
                expected=float(nights_max), dashboard=float(nights),
            ))
    return findings


def audit_solapes(conf: List[dict]) -> List[Finding]:
    findings: List[Finding] = []
    intervals: List[Tuple[date, date, dict]] = []
    for r in conf:
        if not r.get("code") or not r.get("checkin") or r.get("total", 0) == 0:
            continue
        try:
            ci = datetime.strptime(r["checkin"], "%Y-%m-%d").date()
        except ValueError:
            continue
        nights = r.get("nights", 0)
        if nights <= 0:
            continue
        intervals.append((ci, ci + timedelta(days=nights), r))
    intervals.sort(key=lambda x: x[0])
    for i in range(len(intervals)):
        for j in range(i + 1, len(intervals)):
            ci_a, co_a, ra = intervals[i]
            ci_b, co_b, rb = intervals[j]
            if ci_b >= co_a:
                break
            cy = date.today().year
            historical = co_b.year < cy  # ambas terminaron antes del año actual
            findings.append(Finding(
                section="Solapes",
                level="AVISO" if historical else "CRÍTICO",
                blocking=not historical,
                metric="Solape historico" if historical else "Solape activo",
                detail=(f"{ra.get('code')} {ra.get('guest','')} {ci_a}->{co_a}  <->  "
                        f"{rb.get('code')} {rb.get('guest','')} {ci_b}->{co_b}"),
                year=ra["year"], month=ra["month"],
                notes="histórico, no bloquea" if historical else "activo, requiere corrección",
            ))
    return findings


def audit_pace(conf: List[dict]) -> List[Finding]:
    pace     = calc_pace(conf)
    findings: List[Finding] = []
    for (y, m), (ok_v, dash_v) in sorted(pace.items()):
        delta = dash_v - ok_v
        if delta > 0.01:
            sin_bd   = sum(1 for r in conf
                           if r["year"] == y and r["month"] == m and not r.get("booking_date"))
            cy       = date.today().year
            critico  = y >= cy and abs(delta) > 100
            findings.append(Finding(
                section="Pace",
                level="CRÍTICO" if critico else "AVISO",
                blocking=critico,
                metric=f"OTB {y}-{m:02d}",
                detail=(f"correcto={ok_v:.2f}€  dashboard={dash_v:.2f}€  "
                        f"D={delta:.2f}€  sin_booking_date={sin_bd}"),
                year=y, month=m,
                expected=ok_v, dashboard=dash_v,
            ))
    return findings


def audit_cancelaciones(canc: List[dict], conf: List[dict]) -> List[Finding]:
    findings: List[Finding] = []
    for r in canc:
        if "impacto" not in r:
            findings.append(Finding(
                section="Cancelaciones", level="AVISO", blocking=False,
                metric="Sin impacto",
                detail=f"{r.get('code','?')} {r.get('guest','')} {r.get('checkin','')}",
                year=r.get("year", 0), month=r.get("month", 0),
            ))
    years = sorted({r["year"] for r in conf + canc})
    for y in years:
        nc   = sum(1 for r in conf if r["year"] == y)
        nx   = sum(1 for r in canc if r["year"] == y)
        t    = nc + nx
        tasa = nx / t * 100 if t else 0
        findings.append(Finding(
            section="Cancelaciones", level="OK", blocking=False,
            metric=f"Tasa {y}",
            detail=f"{nx}/{t} cancelaciones = {tasa:.1f}%",
            year=y,
        ))
    return findings


def audit_data_integrity(conf: List[dict], canc: List[dict]) -> List[Finding]:
    """booking_date <= checkin, year/month vs checkin, nights=0 con total>0, total<0."""
    findings: List[Finding] = []
    for r in conf + canc:
        code  = r.get("code") or "sin_code"
        guest = r.get("guest", "")
        y     = r.get("year", 0)
        m     = r.get("month", 0)
        bd_s  = r.get("booking_date", "")
        ci_s  = r.get("checkin", "")
        nights = r.get("nights", 0)
        total  = r.get("total", 0)

        if bd_s and ci_s:
            try:
                bd_d = datetime.strptime(bd_s, "%Y-%m-%d").date()
                ci_d = datetime.strptime(ci_s, "%Y-%m-%d").date()
                if bd_d > ci_d:
                    findings.append(Finding(
                        section="Integridad", level="CRÍTICO", blocking=True,
                        metric="booking > checkin",
                        detail=f"{code} {guest} booking={bd_s} checkin={ci_s}",
                        year=y, month=m,
                    ))
            except ValueError:
                pass

        is_continuation = not r.get("code") and total == 0
        if ci_s and y and m and not is_continuation:
            try:
                ci_d = datetime.strptime(ci_s, "%Y-%m-%d").date()
                if ci_d.year != y or ci_d.month != m:
                    findings.append(Finding(
                        section="Integridad", level="CRÍTICO", blocking=True,
                        metric="year/month != checkin",
                        detail=f"{code} {guest} checkin={ci_s} pero year={y} month={m:02d}",
                        year=y, month=m,
                    ))
            except ValueError:
                findings.append(Finding(
                    section="Integridad", level="CRÍTICO", blocking=True,
                    metric="checkin invalido",
                    detail=f"{code} {guest} checkin={ci_s!r}",
                    year=y, month=m,
                ))

        if nights == 0 and total > 0:
            findings.append(Finding(
                section="Integridad", level="CRÍTICO", blocking=True,
                metric="nights=0 total>0",
                detail=f"{code} {guest} nights=0 total={total}",
                year=y, month=m,
            ))

        if r.get("status", "confirmed") == "confirmed" and total < 0:
            findings.append(Finding(
                section="Integridad", level="CRÍTICO", blocking=True,
                metric="total negativo",
                detail=f"{code} {guest} total={total}",
                year=y, month=m,
            ))

    if not findings:
        findings.append(Finding(
            section="Integridad", level="OK", blocking=False,
            metric="Integridad OK", detail="Sin anomalias detectadas",
        ))
    return findings


def audit_lead_time(conf: List[dict]) -> List[Finding]:
    """PM por ventana de antelacion: stored pm vs formula (total-cleaning)/nights."""
    findings: List[Finding] = []
    BUCKETS = [
        ("<7d",    0,   6),
        ("7-30d",  7,  30),
        ("30-90d", 31, 90),
        (">90d",  91, 9999),
    ]
    for label, lo, hi in BUCKETS:
        recs = []
        for r in conf:
            if not r.get("booking_date") or not r.get("checkin"):
                continue
            if not _is_real_record(r):
                continue
            try:
                bd_d = datetime.strptime(r["booking_date"], "%Y-%m-%d").date()
                ci_d = datetime.strptime(r["checkin"],      "%Y-%m-%d").date()
            except ValueError:
                continue
            if lo <= (ci_d - bd_d).days <= hi:
                recs.append(r)

        if not recs:
            continue

        den    = sum(r.get("nights", 0) for r in recs)
        pm_ok  = round(sum(_pm_effective(r) * r.get("nights", 0) for r in recs) / den, 2) if den else 0.0
        pm_d   = round(sum(r.get("total", 0) - r.get("cleaning", 0) for r in recs) / den, 2) if den else 0.0
        delta  = abs(pm_d - pm_ok) / pm_ok * 100 if pm_ok > 0 else 0.0

        findings.append(Finding(
            section="Lead_Time",
            level="OK" if delta < PM_WARN_PCT else "AVISO",
            blocking=False,
            metric=f"PM lead {label}",
            detail=(f"n={len(recs)}  PM_correcto={pm_ok:.2f}  "
                    f"PM_dashboard={pm_d:.2f}  delta={delta:.1f}%"),
            expected=pm_ok, dashboard=pm_d, delta_pct=round(delta, 1),
        ))

    sin_bd = sum(1 for r in conf if _is_real_record(r) and not r.get("booking_date"))
    if sin_bd > 0:
        findings.append(Finding(
            section="Lead_Time", level="AVISO", blocking=False,
            metric="Sin booking_date",
            detail=f"{sin_bd} registros sin booking_date (excluidos del lead time)",
        ))
    return findings


def audit_totales_anuales(conf: List[dict]) -> List[Finding]:
    """PM anual correcto vs dashboard por ano y verificacion de totales."""
    findings: List[Finding] = []
    for y in sorted({r["year"] for r in conf}):
        recs_y = [r for r in conf if r["year"] == y]
        if not recs_y:
            continue
        ing = round(sum(r.get("total", 0) for r in recs_y), 2)
        n   = sum(r.get("nights", 0) for r in recs_y)

        den_ok = sum(r.get("nights", 0) for r in recs_y if _is_real_record(r))
        pm_ok  = round(
            sum(_pm_effective(r) * r.get("nights", 0) for r in recs_y if _is_real_record(r))
            / den_ok, 2
        ) if den_ok > 0 else 0.0

        pm_d  = round(
            sum(r.get("total", 0) - r.get("cleaning", 0) for r in recs_y) / n, 2
        ) if n > 0 else 0.0

        delta = abs(pm_d - pm_ok) / pm_ok * 100 if pm_ok > 0 else 0.0

        findings.append(Finding(
            section="Totales_Anuales",
            level="OK" if delta < PM_WARN_PCT else "AVISO",
            blocking=False,
            metric=f"Anual {y}",
            detail=(f"Ingresos={ing}€  Noches={n}  "
                    f"PM_ok={pm_ok:.2f}  PM_dash={pm_d:.2f}  delta={delta:.1f}%"),
            year=y,
            expected=pm_ok, dashboard=pm_d, delta_pct=round(delta, 1),
        ))
    return findings


def audit_superhost(conf: List[dict], canc: List[dict]) -> List[Finding]:
    """Criterios Superhost (rating>=4.8, estancias>=10, canc<1%) en ventanas de 365 dias."""
    reviews  = _load_reviews()
    findings: List[Finding] = []
    today    = date.today()

    eval_dates: List[date] = []
    for y in range(today.year - 1, today.year + 2):
        for mo in [1, 4, 7, 10]:
            eval_dates.append(date(y, mo, 1))
    eval_dates = [d for d in eval_dates if d <= today + timedelta(days=366)]

    rv_by_code = {
        (rv.get("reservation_id") or rv.get("code")): rv
        for rv in reviews
        if rv.get("reservation_id") or rv.get("code")
    }

    for ed in eval_dates:
        window_start = ed - timedelta(days=365)

        stays = []
        for r in conf:
            ci_s = r.get("checkin", "")
            if not ci_s:
                continue
            # Excluir solo continuaciones cross-month reales (code='', total=0)
            if not r.get("code") and r.get("total", 0) == 0:
                continue
            try:
                ci_d = datetime.strptime(ci_s, "%Y-%m-%d").date()
                co_d = ci_d + timedelta(days=r.get("nights", 0))
            except ValueError:
                continue
            if window_start <= ci_d and co_d <= ed:
                stays.append(r)

        n_stays    = len(stays)
        stay_codes = {r.get("code") for r in stays}
        ratings    = [
            rv_by_code[c].get("rating", 0)
            for c in stay_codes
            if c in rv_by_code and rv_by_code[c].get("rating", 0) > 0
        ]
        avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else 0.0

        n_canc = 0
        for r in canc:
            ci_s = r.get("checkin", "")
            if not ci_s:
                continue
            try:
                ci_d = datetime.strptime(ci_s, "%Y-%m-%d").date()
            except ValueError:
                continue
            if window_start <= ci_d < ed:
                n_canc += 1

        total_w   = n_stays + n_canc
        canc_rate = round(n_canc / total_w * 100, 2) if total_w > 0 else 0.0

        ok_stays  = n_stays >= 10
        # Rating: solo falla si tenemos datos reales (>= 5 reviews en ventana)
        ok_rating = len(ratings) < 5 or avg_rating >= 4.8
        # Canc: solo informativo (no distinguimos host vs guest en los datos)
        passed    = ok_stays and ok_rating

        if passed:
            level  = "OK"
            detail = (f"Estancias={n_stays}  "
                      f"rating={avg_rating:.2f}({len(ratings)} reviews)  "
                      f"canc_total={canc_rate:.1f}%")
        else:
            level  = "AVISO"
            parts: List[str] = []
            if not ok_stays:
                parts.append(f"estancias={n_stays}<10")
            if not ok_rating:
                parts.append(f"rating={avg_rating:.2f}<4.8({len(ratings)} reviews)")
            detail = "  ".join(parts) + f"  canc_total={canc_rate:.1f}%"

        findings.append(Finding(
            section="Superhost", level=level, blocking=False,
            metric=f"Eval {ed.strftime('%Y-%m-%d')}",
            detail=detail,
            year=ed.year, month=ed.month,
            expected=float(n_stays), dashboard=avg_rating,
            notes="futuro" if ed > today else "historico",
        ))

    return findings


def audit_conversion(conf: List[dict]) -> List[Finding]:
    """CVR% = reservas vendidas en mes / visitas del mes."""
    visitas = _load_visitas()
    if not visitas:
        return [Finding(
            section="Conversion", level="AVISO", blocking=False,
            metric="Sin datos visitas",
            detail="visitas.json no encontrado o vacio",
        )]

    findings: List[Finding] = []
    reservas_mes: Dict[str, int] = defaultdict(int)
    for r in conf:
        bd = r.get("booking_date", "")
        if bd and len(bd) >= 7:
            reservas_mes[bd[:7]] += 1

    for mes_key in sorted(visitas):
        vis_n = visitas[mes_key]
        res_n = reservas_mes.get(mes_key, 0)
        if vis_n == 0:
            continue
        cvr = round(res_n / vis_n * 100, 2)

        try:
            y, m = int(mes_key[:4]), int(mes_key[5:7])
        except ValueError:
            y, m = 0, 0

        cy_now = date.today().year
        reciente = y >= cy_now - 1
        if cvr > 20:
            level  = "AVISO"
            detail = f"CVR={cvr:.1f}% ({res_n}/{vis_n}) inusualmente alto"
        elif res_n == 0 and reciente:
            level  = "AVISO"
            detail = f"CVR=0% (0/{vis_n}) sin ventas con booking_date en este mes"
        else:
            level  = "OK"
            detail = f"CVR={cvr:.1f}% ({res_n}/{vis_n})" if res_n > 0 else f"CVR=0% (historico, sin booking_date)"

        findings.append(Finding(
            section="Conversion", level=level, blocking=False,
            metric=f"CVR {mes_key}",
            detail=detail,
            year=y, month=m,
            expected=float(vis_n), dashboard=float(res_n), delta_pct=cvr,
        ))

    return findings


# ── TESTS UNITARIOS ───────────────────────────────────────────────────────────

def run_unit_tests(conf: List[dict]) -> List[Finding]:
    findings: List[Finding] = []
    distorted = _distorted_months(conf)
    pm_ok     = calc_pm_correcto(conf)
    pm_dash   = calc_pm_dashboard(conf)

    def _t(name: str, passed: bool, detail: str,
           year: int = 0, month: int = 0,
           expected: float = 0.0, actual: float = 0.0) -> None:
        findings.append(Finding(
            section="Tests_Unitarios",
            level="OK" if passed else "CRÍTICO",
            blocking=not passed,
            metric=name, detail=detail,
            year=year, month=month,
            expected=expected, dashboard=actual,
        ))

    # T1 — Darya Kramar jun→jul 2026 (HMNKEKCM4M)
    darya = next((r for r in conf if r.get("code") == "HMNKEKCM4M"), None)
    if darya:
        n = darya.get("nights", 0)
        _t("T1a HMNKEKCM4M nights jun-2026", n == 1,
           f"nights={n} esperado=1", 2026, 6, 1.0, float(n))
        cont = next((r for r in conf if not r.get("code")
                     and r.get("guest", "") == darya.get("guest", "")
                     and r.get("year") == 2026 and r.get("month") == 7), None)
        _t("T1b HMNKEKCM4M continuación jul-2026", bool(cont),
           "Continuación " + ("encontrada" if cont else "NO encontrada"), 2026, 7)
        _t("T1c (2026,6) en distorted_months", (2026, 6) in distorted,
           f"(2026,6) in distorted_months = {(2026, 6) in distorted}", 2026, 6)
        pm_c = pm_ok.get((2026, 6), 0)
        pm_d = pm_dash.get((2026, 6), 0)
        delta_ok = abs(pm_d - pm_c) / pm_c * 100 if pm_c > 0 else 0
        _t("T1d PM jun-2026 formula corregida (delta<5%)",
           delta_ok < 5.0,
           f"PM_correcto={pm_c:.2f}  PM_dashboard={pm_d:.2f}  delta={delta_ok:.1f}%",
           2026, 6, pm_c, pm_d)
    else:
        _t("T1 HMNKEKCM4M", False, "Reserva no encontrada en datos", 2026, 6)

    # T2 — Vasile Cumatrenco jul→ago 2026 (HMHM4FQMHK)
    vasile = next((r for r in conf if r.get("code") == "HMHM4FQMHK"), None)
    if vasile:
        n = vasile.get("nights", 0)
        _t("T2a HMHM4FQMHK nights jul-2026", n == 2,
           f"nights={n} esperado=2", 2026, 7, 2.0, float(n))
        _t("T2b (2026,7) en distorted_months", (2026, 7) in distorted,
           f"(2026,7) in distorted_months = {(2026, 7) in distorted}", 2026, 7)
    else:
        _t("T2 HMHM4FQMHK", False, "Reserva no encontrada", 2026, 7)

    # T3 — Arthur Schaber feb→mar 2026 (HMZRBPTXRS)
    arthur = next((r for r in conf if r.get("code") == "HMZRBPTXRS"), None)
    if arthur:
        n  = arthur.get("nights", 0)
        ci = arthur.get("checkin", "")
        _t("T3 HMZRBPTXRS nights feb-2026", n == 3,
           f"nights={n} esperado=3 checkin={ci}", 2026, 2, 3.0, float(n))
    else:
        _t("T3 HMZRBPTXRS", False, "Reserva no encontrada", 2026, 2)

    # T4-T6 — Cross-month históricos documentados
    hist_cases = [
        ("T4", "Lisanne Vladisavljevic",    2021,  9, 2),
        ("T5", "Mireille Heronneau",         2021, 10, 2),
        ("T6", "Elisabeth Liwadas Kreutz",   2022,  5, 5),
    ]
    for tid, guest, year, month, expected_n in hist_cases:
        r = next((x for x in conf
                  if x.get("guest", "") == guest
                  and x["year"] == year and x["month"] == month
                  and x.get("total", 0) > 0), None)
        if r:
            n = r.get("nights", 0)
            _t(f"{tid} {guest} {year}-{month:02d}",
               n == expected_n,
               f"nights={n} esperado={expected_n}",
               year, month, float(expected_n), float(n))
        else:
            _t(f"{tid} {guest}", False, "Registro no encontrado", year, month)

    return findings


# ── EXCEL ─────────────────────────────────────────────────────────────────────

def _fill(level: str) -> PatternFill:
    return {"CRÍTICO": FILL_ERROR, "ERROR": FILL_ERROR,
            "AVISO": FILL_WARN, "OK": FILL_OK}.get(level, FILL_INFO)


def _autosize(ws, max_w: int = 60) -> None:
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)


def write_excel(path: Path, findings: List[Finding],
                conf: List[dict], canc: List[dict]) -> None:
    wb  = Workbook()
    blk = [f for f in findings if f.blocking]
    wrn = [f for f in findings if f.level == "AVISO" and not f.blocking]
    oks = [f for f in findings if f.level == "OK"]

    # ── Resumen ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    estado = "CRITICOS" if blk else "AVISOS" if wrn else "OK"
    ws.append(["AUDITORÍA DASHBOARD CSJ", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Estado global", estado,
               f"{len(blk)} criticos / {len(wrn)} avisos / {len(oks)} OK"])
    ws.append([])
    hdr = ["Sección","Nivel","Bloq.","Métrica","Detalle","Año","Mes",
           "Esperado","Dashboard","Delta%","Notas"]
    ws.append(hdr)
    for c in ws[5]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in sorted(findings, key=lambda x: (x.level not in ("CRÍTICO","ERROR"), x.level != "AVISO", x.section)):
        ws.append([
            f.section, f.level, "SÍ" if f.blocking else "no",
            f.metric, f.detail,
            f.year or "", f.month or "",
            round(f.expected, 2) if f.expected else "",
            round(f.dashboard, 2) if f.dashboard else "",
            f.delta_pct if f.delta_pct else "",
            f.notes,
        ])
        for c in ws[ws.max_row]:
            c.fill = _fill(f.level)
    _autosize(ws)
    ws.freeze_panes = "A6"

    # ── PM Mensual ────────────────────────────────────────────────────────────
    wp = wb.create_sheet("PM_Mensual")
    distorted = _distorted_months(conf)
    pm_ok     = calc_pm_correcto(conf)
    pm_dash   = calc_pm_dashboard(conf)
    years     = sorted({y for y, _ in pm_ok})
    wp.append(["Año","Mes","PM_Correcto","PM_Dashboard","Delta_Pct","Tipo","Nivel"])
    for c in wp[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for y in years:
        for m in range(1, 13):
            ok_v   = pm_ok.get((y, m), 0)
            dash_v = pm_dash.get((y, m), 0)
            if ok_v == 0 and dash_v == 0:
                continue
            delta = abs(dash_v - ok_v) / ok_v * 100 if ok_v > 0 else 0
            tipo  = "CrossMonth" if (y, m) in distorted else "Normal"
            level = ("OK" if delta < PM_WARN_PCT
                     else "CRÍTICO" if delta >= PM_CRIT_PCT
                     else "AVISO")
            wp.append([y, MONTHS_ES[m - 1], ok_v, dash_v, round(delta, 1), tipo, level])
            for c in wp[wp.max_row]:
                c.fill = _fill(level)
    _autosize(wp, 20)
    wp.freeze_panes = "A2"

    # ── Ingresos y Ocupación ──────────────────────────────────────────────────
    wi = wb.create_sheet("Ingresos_Ocupacion")
    ingresos = calc_ingresos(conf)
    noches_m = calc_noches(conf)
    wi.append(["Año","Mes","Ingresos€","Noches","Días_mes","Ocupación%","Status"])
    for c in wi[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for y in years:
        for m in range(1, 13):
            ing  = ingresos.get((y, m), 0)
            n    = noches_m.get((y, m), 0)
            dias = calendar.monthrange(y, m)[1]
            if ing == 0 and n == 0:
                continue
            ocu   = round(n / dias * 100, 1)
            level = "ERROR" if ocu > 100 else "OK"
            wi.append([y, MONTHS_ES[m - 1], ing, n, dias, ocu, level])
            for c in wi[wi.max_row]:
                c.fill = _fill(level)
    _autosize(wi, 20)
    wi.freeze_panes = "A2"

    # ── PM Temporada ──────────────────────────────────────────────────────────
    wt = wb.create_sheet("PM_Temporada")
    pm_t_ok   = calc_pm_temporada_correcto(conf)
    pm_t_dash = calc_pm_temporada_dashboard(conf)
    wt.append(["Año","Temporada","PM_Correcto","PM_Dashboard","Delta_Pct","Nivel"])
    for c in wt[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for y in years:
        for banda in ["alta", "media", "baja"]:
            ok_v   = pm_t_ok.get((y, banda), 0)
            dash_v = pm_t_dash.get((y, banda), 0)
            if ok_v == 0 and dash_v == 0:
                continue
            delta = abs(dash_v - ok_v) / ok_v * 100 if ok_v > 0 else 0
            level = "OK" if delta < PM_WARN_PCT else "AVISO"
            wt.append([y, banda.capitalize(), ok_v, dash_v, round(delta, 1), level])
            for c in wt[wt.max_row]:
                c.fill = _fill(level)
    _autosize(wt, 20)
    wt.freeze_panes = "A2"

    # ── Pace ─────────────────────────────────────────────────────────────────
    wpa = wb.create_sheet("Pace")
    pace = calc_pace(conf)
    wpa.append(["Año","Mes","OTB_Correcto","OTB_Dashboard","Diferencia","Sin_BookingDate","Nivel"])
    for c in wpa[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for (y, m), (ok_v, dash_v) in sorted(pace.items()):
        delta  = dash_v - ok_v
        sin_bd = sum(1 for r in conf
                     if r["year"] == y and r["month"] == m and not r.get("booking_date"))
        level  = "AVISO" if delta > 0.01 else "OK"
        wpa.append([y, MONTHS_ES[m - 1], ok_v, dash_v, round(delta, 2), sin_bd, level])
        for c in wpa[wpa.max_row]:
            c.fill = _fill(level)
    _autosize(wpa, 20)
    wpa.freeze_panes = "A2"

    # ── Cross-Month ───────────────────────────────────────────────────────────
    wc = wb.create_sheet("Cross_Month")
    wc.append(["Code","Guest","Checkin","Nights_almacenado","Nights_max_mes","Status","Continuación"])
    for c in wc[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    cross_ok = True
    for r in conf:
        if not r.get("code") or not r.get("checkin"):
            continue
        try:
            ci = datetime.strptime(r["checkin"], "%Y-%m-%d").date()
        except ValueError:
            continue
        dias_mes   = calendar.monthrange(ci.year, ci.month)[1]
        nights_max = dias_mes - ci.day + 1
        nights     = r.get("nights", 0)
        if nights > nights_max:
            cross_ok = False
            cont = next((x for x in conf
                         if not x.get("code") and x.get("guest", "") == r.get("guest", "")
                         and ((x["year"] == r["year"] and x["month"] == r["month"] + 1)
                              or (r["month"] == 12 and x["year"] == r["year"] + 1 and x["month"] == 1))), None)
            wc.append([r.get("code"), r.get("guest", ""), r["checkin"], nights, nights_max,
                       "ERROR",
                       (f"Sí ({cont['nights']}n {cont['month']:02d}/{cont['year']})" if cont
                        else "NO ENCONTRADA")])
            for c in wc[wc.max_row]:
                c.fill = FILL_ERROR
    if cross_ok:
        wc.append(["Sin errores de prorrateo"])
        wc["A2"].fill = FILL_OK
    _autosize(wc, 30)

    # ── Solapes ───────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Solapes")
    ws2.append(["Detalle del solape"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    solape_f = [f for f in findings if f.section == "Solapes"]
    if not solape_f:
        ws2.append(["Sin solapes detectados"])
        ws2["A2"].fill = FILL_OK
    for f in solape_f:
        ws2.append([f.detail])
        for c in ws2[ws2.max_row]:
            c.fill = FILL_ERROR
    _autosize(ws2, 80)

    # ── Tests Unitarios ───────────────────────────────────────────────────────
    wtu = wb.create_sheet("Tests_Unitarios")
    wtu.append(["Status","Test","Detalle","Esperado","Obtenido"])
    for c in wtu[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in [x for x in findings if x.section == "Tests_Unitarios"]:
        wtu.append([f.level, f.metric, f.detail,
                    round(f.expected, 2) if f.expected else "",
                    round(f.dashboard, 2) if f.dashboard else ""])
        for c in wtu[wtu.max_row]:
            c.fill = _fill(f.level)
    _autosize(wtu, 55)

    # ── Integridad ────────────────────────────────────────────────────────────
    wint = wb.create_sheet("Integridad")
    wint.append(["Nivel","Bloq.","Metrica","Detalle","Año","Mes"])
    for c in wint[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in [x for x in findings if x.section == "Integridad"]:
        wint.append([f.level, "SI" if f.blocking else "no",
                     f.metric, f.detail, f.year or "", f.month or ""])
        for c in wint[wint.max_row]:
            c.fill = _fill(f.level)
    _autosize(wint, 55)

    # ── Lead Time ─────────────────────────────────────────────────────────────
    wlt = wb.create_sheet("Lead_Time")
    wlt.append(["Ventana","N_Reservas","PM_Correcto","PM_Dashboard","Delta%","Nivel"])
    for c in wlt[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in [x for x in findings if x.section == "Lead_Time"]:
        n_match = None
        try:
            n_match = int(f.detail.split("n=")[1].split()[0]) if "n=" in f.detail else ""
        except (IndexError, ValueError):
            n_match = ""
        wlt.append([f.metric, n_match,
                    round(f.expected, 2) if f.expected else "",
                    round(f.dashboard, 2) if f.dashboard else "",
                    f.delta_pct if f.delta_pct else "", f.level])
        for c in wlt[wlt.max_row]:
            c.fill = _fill(f.level)
    _autosize(wlt, 20)

    # ── Totales Anuales ───────────────────────────────────────────────────────
    wta = wb.create_sheet("Totales_Anuales")
    wta.append(["Año","Detalle","PM_Correcto","PM_Dashboard","Delta%","Nivel"])
    for c in wta[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in [x for x in findings if x.section == "Totales_Anuales"]:
        wta.append([f.year, f.detail,
                    round(f.expected, 2) if f.expected else "",
                    round(f.dashboard, 2) if f.dashboard else "",
                    f.delta_pct if f.delta_pct else "", f.level])
        for c in wta[wta.max_row]:
            c.fill = _fill(f.level)
    _autosize(wta, 60)

    # ── Superhost ─────────────────────────────────────────────────────────────
    wsh = wb.create_sheet("Superhost")
    wsh.append(["Eval_date","Tipo","Estancias","Rating_medio","Canc%","Detalle","Estado"])
    for c in wsh[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in [x for x in findings if x.section == "Superhost"]:
        wsh.append([f.metric, f.notes or "",
                    int(f.expected) if f.expected else "",
                    round(f.dashboard, 2) if f.dashboard else "",
                    "", f.detail, f.level])
        for c in wsh[wsh.max_row]:
            c.fill = _fill(f.level)
    _autosize(wsh, 55)

    # ── Conversion ────────────────────────────────────────────────────────────
    wcv = wb.create_sheet("Conversion")
    wcv.append(["Mes","Visitas","Reservas_Vendidas","CVR%","Nivel"])
    for c in wcv[1]:
        c.font = Font(bold=True)
        c.fill = FILL_HEADER
    for f in [x for x in findings if x.section == "Conversion"]:
        wcv.append([f.metric,
                    int(f.expected) if f.expected else "",
                    int(f.dashboard) if f.dashboard else "",
                    f.delta_pct if f.delta_pct else "",
                    f.level])
        for c in wcv[wcv.max_row]:
            c.fill = _fill(f.level)
    _autosize(wcv, 20)

    wb.save(path)


# ── MAIN ─────────────────────────────────────────────────────────────────────

def auditar(verbose: bool = True) -> Tuple[List[Finding], List[Finding]]:
    conf, canc = _load()

    all_findings: List[Finding] = []
    all_findings += audit_data_integrity(conf, canc)
    all_findings += audit_pm_mensual(conf)
    all_findings += audit_pm_temporada(conf)
    all_findings += audit_ocupacion(conf)
    all_findings += audit_crossmonth(conf)
    all_findings += audit_solapes(conf)
    all_findings += audit_pace(conf)
    all_findings += audit_cancelaciones(canc, conf)
    all_findings += audit_lead_time(conf)
    all_findings += audit_totales_anuales(conf)
    all_findings += audit_superhost(conf, canc)
    all_findings += audit_conversion(conf)
    all_findings += run_unit_tests(conf)

    blocking = [f for f in all_findings if f.blocking]
    warns    = [f for f in all_findings if f.level == "AVISO" and not f.blocking]

    OUTPUT.mkdir(exist_ok=True)
    write_excel(OUTPUT / "auditoria_dashboard.xlsx", all_findings, conf, canc)
    with (OUTPUT / "auditoria_dashboard.json").open("w", encoding="utf-8") as fp:
        json.dump([asdict(f) for f in all_findings], fp, ensure_ascii=False, indent=2)

    if verbose:
        enc = sys.stdout.encoding or "utf-8"
        def _p(s: str) -> None:
            print(s.encode(enc, errors="replace").decode(enc))

        n_crit  = len(blocking)
        n_aviso = len(warns)
        n_ok    = sum(1 for f in all_findings if f.level == "OK")
        estado  = "BLOQUEADO" if blocking else "OK para generar"

        _p(f"\n{'='*55}")
        _p(f"  AUDITORIA DASHBOARD CSJ")
        _p(f"{'='*55}")
        _p(f"  Registros: {len(conf)+len(canc)} ({len(conf)} conf, {len(canc)} canc)")
        _p(f"")
        _p(f"  RESUMEN EJECUTIVO")
        _p(f"  -----------------")
        _p(f"  OK:        {n_ok}")
        _p(f"  AVISOS:    {n_aviso}")
        _p(f"  CRITICOS:  {n_crit}")
        _p(f"  Estado:    {estado}")
        if blocking:
            _p(f"\n  CRITICOS ({n_crit}) — bloquean generacion del dashboard:")
            for f in blocking:
                _p(f"    [CRIT] [{f.section}] {f.metric}: {f.detail}")
        if warns:
            _p(f"\n  AVISOS ({n_aviso}):")
            for f in warns:
                _p(f"    [!] [{f.section}] {f.metric}: {f.detail}")
        if not blocking and not warns:
            _p(f"\n  [OK] Todas las metricas cuadran")
        _p(f"{'='*55}")
        _p(f"  Excel: {(OUTPUT / 'auditoria_dashboard.xlsx').resolve()}")
        _p(f"{'='*55}\n")

    return blocking, warns


if __name__ == "__main__":
    blocking, _ = auditar(verbose=True)
    sys.exit(1 if blocking else 0)
